#!/usr/bin/env python3
"""
One-time, READ-ONLY mining script for the Expenses domain (Phase 1 — Foundation).

CANONICAL CONTRACT v2 (matches Agent B's TS ingest exactly):
  - Ingest operates on the RAW Cyrillic bank rows, so all seed data is mined from
    the RAW sources (not the romanized _english files, which caused the v1
    2.7% vendor / 12.3% dictionary hit-rate failure).
  - Translation dictionary KEY = normalizeDescription(`Основание за плащане` + ' '
    + `Описание на операцията`), in that order, EXCLUDING `Още пояснения`.
  - Vendor match strings are the RAW Cyrillic counterparty field (`Бенефициент`
    for outgoing/debit rows), matched case-insensitively, match_type 'exact'.
  - normalizeDescription digit regex is ASCII-only ([0-9]) to match JS \\d parity.

RAW sources (all READ-ONLY):
  - report(3).xls               HTML table (parsed with stdlib html.parser)
  - report(1).xls.xlsx          real OOXML, Cyrillic headers (openpyxl)
  - report(2).xls.xlsx          real OOXML, English headers / Cyrillic data (openpyxl)
                                — card-statement style, NO counterparty columns.
Each raw row is aligned to its report(N)_english.xlsx `Transactions` row (which
carries Category and the romanized `Description Original`) by
(reference + amount + value_date), with a positional fallback.

Emits (idempotent):
  supabase/migrations/127_expense_mapping_seed.sql
  scripts/expenses/output/{translation_dict,vendor_rules,keyword_rules}.{csv,json}
  scripts/expenses/output/{overhead_by_category,mining_stats}.json

No network, no AI, no writes into the reference folder.
"""

import csv
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from html.parser import HTMLParser

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

CANDIDATE_REFERENCE_DIRS = [
    "/mnt/c/Users/Matthew/Dropbox/Organizations/The B Team/Financials/2025-AP",
    "/mnt/c/Users/Matthew/Documents/Dropbox/Organizations/The B Team/Financials/2025-AP",
]

REPORTS = [1, 2, 3]
PREFERRED_REPORT = 3

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(SCRIPT_DIR, "output")
MIGRATIONS_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "supabase", "migrations"))
SEED_SQL_PATH = os.path.join(MIGRATIONS_DIR, "127_expense_mapping_seed.sql")

# Frozen category set — MUST match migration 125 fixed ids exactly.
CATEGORY_IDS = {
    "Payroll": 1, "Payroll Taxes": 2, "Software & AI Tools": 3,
    "Bank & Transfer Fees": 4, "Vehicle & Mobility": 5, "Office Supplies & Food": 6,
    "Contractors & Agency Fees": 7, "Employee Benefits": 8, "Office Operations": 9,
    "Utilities & Facilities": 10, "Telecom & Internet": 11, "Treasury & Wallet Transfers": 12,
    "Accounting & Compliance": 13, "Debt Service": 14, "Miscellaneous": 15,
}
FALLBACK_CATEGORY_ID = 15
ID_TO_NAME = {v: k for k, v in CATEGORY_IDS.items()}

VENDOR_PRIORITY_RAW = 100        # primary: RAW Cyrillic beneficiary rules
VENDOR_PRIORITY_ROMANIZED = 200  # secondary: romanized english Vendor rules (future xlsx uploads)

# Keyword candidates re-validated against RAW canonical description text.
# (keyword, category_id, priority). Latin merchant names survive in raw POS lines;
# Bulgarian/Cyrillic terms cover the bank-transfer rows. The validator drops any
# that fail (>=3 hits, >=90% dominance, and never resolving to category 15).
KEYWORD_CANDIDATES = [
    # Latin merchant tokens (appear verbatim in raw POS "Основание" text) -> Software
    ("CURSOR", 3, 10), ("GITHUB", 3, 10), ("GODADDY", 3, 10), ("CLAUDE.AI", 3, 10),
    ("OPENAI", 3, 10), ("CHATGPT", 3, 10), ("FIGMA", 3, 10), ("MICROSOFT", 3, 10),
    ("BROWSERSTACK", 3, 10), ("TWILIO", 3, 10), ("VERCEL", 3, 10), ("NETLIFY", 3, 10),
    # Contractors / agency
    ("AGENT SERVIC", 7, 20), ("MAIN ASSEMBLY", 7, 20),
    # Bank & transfer fees — Cyrillic phrases as they appear in raw
    ("ТАКСИ ПОЛУЧЕНИ ВАЛУТНИ", 4, 20), ("ТАКСИ ПЛАТЕНИ ВАЛУТНИ", 4, 20),
    ("ТАКСА ПРЕВОД", 4, 20), ("RECURRING FEE", 4, 20), ("AZV-COMMISSION", 4, 20),
    # Payroll — both scripts
    ("ЗАПЛАТА", 1, 30), ("SALARY", 1, 30),
    # Payroll taxes / social security  (Payroll Taxes is otherwise covered by the
    # NAP counterparty vendor rule; short abbreviations like "ДОД" are excluded as
    # unsafe substrings.)
    ("ОСИГУРОВКИ", 2, 25), ("ОСИГУРОВК", 2, 25),
    # Telecom
    ("A1 БЪЛГАРИЯ", 11, 20), ("ТЕЛЕНОР", 11, 20), ("VIVACOM", 11, 20),
    # Vehicle & mobility
    ("ГОРИВО", 5, 25), ("ЛИЗИНГ", 5, 25), ("ПАРКИНГ", 5, 25),
    # Utilities / facilities
    ("ЕЛЕКТРО", 10, 25), ("ТОПЛОФИКАЦИЯ", 10, 25), ("ВОДОСНАБДЯВАНЕ", 10, 25),
]
MIN_KEYWORD_HITS = 3
MIN_KEYWORD_DOMINANCE = 0.90
OVERHEAD_DOMINANCE = 0.90

# Keyword rules that MATCH but must flag their matches for human review at ingest.
# UNICREDIT BULBANK is the bank's own name — a broad substring that legitimately
# maps to Debt Service in the reference, but risks silently misfiling unrelated
# bank-name mentions. The rule stays; every match is forced to needs_review until
# the user confirms/narrows it. (Agent B honors the exact column name force_review.)
FORCE_REVIEW_KEYWORDS = {"UNICREDIT BULBANK", "UNICREDIT BULBANK AD"}

# ---------------------------------------------------------------------------
# DELTA (migration 128) — additive keyword hardening. 127 is ALREADY APPLIED to
# prod and MUST NOT be regenerated; run `mine_expenses.py --delta` to emit ONLY
# the new rules (idempotent ON CONFLICT). Root cause these fix: the seed derives
# merchant keywords from the *clean romanized Vendor name* (e.g. "HIGHLEVEL INC.",
# "KAFFEKAPSLEN", "EPAY OFFICE1.BG"), but real card/POS description lines carry a
# shorter merchant *descriptor* variant ("HIGHLEVEL AGENCY SUB", "KaffeK/...",
# "office1.bg") that the longer seeded keyword is not a substring of — so the row
# falls through to Miscellaneous even though the same vendor is categorized
# elsewhere. Each token below is a shorter, still-unambiguous substring of that
# vendor's POS variant. All are re-validated against the reference corpus by the
# SAME validator as the seed (>=3 hits, >=90% dominance, non-15 dominant).
DELTA_KEYWORD_CANDIDATES = [
    ("KAFFEK", 6, 15),       # Office Supplies & Food — POS "KaffeK/Hasselager/DNK"
    ("HIGHLEVEL", 3, 15),    # Software & AI Tools — POS "HIGHLEVEL AGENCY SUB", "HIGHLEVEL * TRIAL OVER"
    ("OFFICE1.BG", 9, 15),   # Office Operations — POS "office1.bg" (seed had only "EPAY OFFICE1.BG")
    # --- User-mandated (business owner), each independently reference-backed ---
    # These match the CYRILLIC description_original (what categorize() actually
    # reads); they are the effective form of the owner's English directives. All
    # are anchored on a fee-specific prefix so they cannot over-match: the bare
    # "ИЗХ.ПРЕВОД SEPA" marker also appears on salary/tax/insurance transfers
    # (145 already-categorized rows) — only rows PREFIXED with "ТАКСА ЗА" (fee
    # for) are the bank fees.
    ("ТАКСА ЗА ИЗХ.ПРЕВОД SEPA", 4, 20),  # Bank & Transfer Fees — "TAKSA ZA OUTGOING SEPA TRANSFER"; ref 23/23
    ("ПЕРИОДИЧНА ТАКСА", 4, 20),          # Bank & Transfer Fees — "RECURRING FEE DUE" (Дължима периодична такса); ref 25/25
    ("DROPBOX", 3, 10),                    # Software & AI Tools (owner: office software = Software, NOT Office Supplies); ref 79/79. Already seeded in 127 — idempotent no-op re-affirming the mapping.
]

# Documented exceptions to the >=3-reference-hit bar. The reference corpus never
# contained the POS *domain* form of these merchants (their bank rows were
# card/POS lines with no counterparty column in report(2), and report(3)'s HTML
# romanized the vendor to its legal name), so the token scores 0 direct hits even
# though the vendor itself is unambiguously and dominantly categorized in the
# reference. Justification is recorded per row for the audit trail.
DELTA_KEYWORD_EXCEPTIONS = [
    # ("A1.BG", 11, 15, justification)  vendor "A1" -> Telecom & Internet 27/27 (100%)
    #   in the reference; POS variant "www.a1.bg" is the same A1 Bulgaria telecom.
    ("A1.BG", 11, 15, "vendor A1 -> Telecom & Internet 27/27 in reference; POS domain variant of seeded 'A1 BULGARIA LTD'"),
    # English-literal forms the owner named. They match the English
    # description_TRANSLATED, but categorize() only reads description_ORIGINAL,
    # which in the current UniCredit English-UI export is still Cyrillic — so
    # these are INERT on current prod (0 reference hits, 0 prod matches, hence 0
    # over-match) and are seeded purely as future-proofing should the bank ever
    # emit fully-Latin narrative text. The Cyrillic anchors above are what
    # actually recategorize today's rows.
    ("TAKSA ZA OUTGOING SEPA TRANSFER", 4, 20, "owner-named English literal; matches description_translated only; inert on current Cyrillic prod (future-proof)"),
    ("RECURRING FEE DUE", 4, 20, "owner-named English literal; matches description_translated only; inert on current Cyrillic prod (future-proof)"),
]

DELTA_SQL_PATH = os.path.join(MIGRATIONS_DIR, "128_expense_mapping_delta.sql")

# ---------------------------------------------------------------------------
# Shared normalization contract (MUST match the ingestion agent's implementation)
# ---------------------------------------------------------------------------

_NUM_RE = re.compile(r"[0-9]+(?:[.,][0-9]+)*")   # ASCII-only (F6 parity with JS \d)
_WS_RE = re.compile(r"\s+")
_CYRILLIC_RE = re.compile(r"[Ѐ-ӿԀ-ԯ]")


def normalize_description(s):
    """NFC -> uppercase -> ASCII digit runs collapsed to '#' -> ws collapse -> trim."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFC", str(s)).upper()
    s = _NUM_RE.sub("#", s)
    s = _WS_RE.sub(" ", s).strip()
    return s


def has_cyrillic(s):
    return bool(_CYRILLIC_RE.search(str(s))) if s is not None else False


def clean(v):
    """Normalize a raw cell to a clean string (strip nbsp, treat 'None' as empty)."""
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    return "" if s.lower() == "none" else s


def iso_date(v):
    """Coerce a raw date to YYYY-MM-DD. Accepts 'DD.MM.YYYY[ hh:mm:ss]' or ISO."""
    s = clean(v)
    if not s:
        return ""
    s = s.split(" ")[0]
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", s)
    if m:
        return s
    return s


def amount_key(v):
    s = clean(v).replace(",", ".")
    try:
        return f"{float(s):.2f}"
    except (ValueError, TypeError):
        return ""


# ---------------------------------------------------------------------------
# Raw source parsers (READ-ONLY)
# ---------------------------------------------------------------------------

class _HTMLTable(HTMLParser):
    def __init__(self):
        super().__init__()
        self.rows = []
        self._row = None
        self._cell = None

    def handle_starttag(self, tag, attrs):
        if tag == "tr":
            self._row = []
        elif tag in ("td", "th"):
            self._cell = []

    def handle_endtag(self, tag):
        if tag == "tr" and self._row is not None:
            self.rows.append(self._row)
            self._row = None
        elif tag in ("td", "th") and self._cell is not None and self._row is not None:
            self._row.append("".join(self._cell).strip())
            self._cell = None

    def handle_data(self, data):
        if self._cell is not None:
            self._cell.append(data)


# Canonical raw-row shape produced by every parser:
#   {reference, value_date, amount, beneficiary, orderer, osnovanie, opisanie, entry_type}

def _rawrow(reference, value_date, amount, beneficiary, orderer, osnovanie, opisanie, entry_type):
    return {
        "reference": clean(reference).replace(" ", ""),
        "value_date": iso_date(value_date),
        "amount": amount_key(amount),
        "beneficiary": clean(beneficiary),
        "orderer": clean(orderer),
        "osnovanie": clean(osnovanie),   # Основание за плащане
        "opisanie": clean(opisanie),     # Описание на операцията
        "entry_type": clean(entry_type),
    }


def parse_report1(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    h = {name: i for i, name in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        g = lambda k: r[h[k]] if k in h and h[k] < len(r) else None
        out.append(_rawrow(g("Референция"), g("Вальор"), g("Сума във валута на сметката"),
                           g("Бенефициент"), g("Наредител"),
                           g("Основание за плащане"), g("Описание на операцията"), g("Тип")))
    return out


def parse_report2(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    h = {name: i for i, name in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        g = lambda k: r[h[k]] if k in h and h[k] < len(r) else None
        # Card-statement export: NO beneficiary/orderer columns.
        out.append(_rawrow(g("Reference"), g("Value date"), g("Amount in currency of the account"),
                           None, None,
                           g("Details of Payment"), g("Description of the operation"), g("Type")))
    return out


def parse_report3_html(path):
    with open(path, encoding="utf-8") as fh:
        parser = _HTMLTable()
        parser.feed(fh.read())
    rows = [r for r in parser.rows if len(r) >= 16]
    h = {name: i for i, name in enumerate(rows[0])}
    out = []
    for r in rows[1:]:
        g = lambda k: r[h[k]] if k in h and h[k] < len(r) else None
        out.append(_rawrow(g("Референция"), g("Вальор"), g("Сума във валута на сметката"),
                           g("Бенефициент"), g("Наредител"),
                           g("Основание за плащане"), g("Описание на операцията"), g("Тип")))
    return out


def load_english(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Transactions"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    h = rows[0]
    return [dict(zip(h, r)) for r in rows[1:]]


# ---------------------------------------------------------------------------
# Alignment: raw row -> english row (Category, romanized Description Original)
# ---------------------------------------------------------------------------

def build_english_index(eng_rows):
    idx = defaultdict(list)
    for e in eng_rows:
        key = (clean(e.get("Reference")).replace(" ", ""),
               amount_key(e.get("Amount Account Currency")),
               iso_date(e.get("Value Date")))
        idx[key].append(e)
    return idx


def align(raw_rows, eng_rows):
    """Return list parallel to raw_rows of matched english dicts (or None), plus stats."""
    idx = build_english_index(eng_rows)
    pools = {k: list(v) for k, v in idx.items()}
    matched = [None] * len(raw_rows)
    by_key = 0
    for i, rr in enumerate(raw_rows):
        key = (rr["reference"], rr["amount"], rr["value_date"])
        pool = pools.get(key)
        if pool:
            matched[i] = pool.pop(0)
            by_key += 1
    by_pos = 0
    if len(raw_rows) == len(eng_rows):
        for i in range(len(raw_rows)):
            if matched[i] is None:
                matched[i] = eng_rows[i]
                by_pos += 1
    stats = {"raw": len(raw_rows), "english": len(eng_rows),
             "by_key": by_key, "by_pos": by_pos,
             "unaligned": sum(1 for m in matched if m is None)}
    return matched, stats


def eng_category(e):
    return clean(e.get("Category")) if e else ""


def eng_romanized(e):
    return clean(e.get("Description Original")) if e else ""


# ---------------------------------------------------------------------------
# Mining
# ---------------------------------------------------------------------------

def mine(reference_dir):
    raw = {
        1: parse_report1(os.path.join(reference_dir, "report(1).xls.xlsx")),
        2: parse_report2(os.path.join(reference_dir, "report(2).xls.xlsx")),
        3: parse_report3_html(os.path.join(reference_dir, "report(3).xls")),
    }
    english = {n: load_english(os.path.join(reference_dir, f"report({n})_english.xlsx")) for n in REPORTS}

    aligned = {}
    align_stats = {}
    for n in REPORTS:
        aligned[n], align_stats[n] = align(raw[n], english[n])

    unknown_categories = Counter()

    def raw_canonical(rr):
        return (rr["osnovanie"] + " " + rr["opisanie"]).strip()

    # ---- (a) Translation dictionary (key from RAW canonical builder) ----
    dict_map = {}
    for n in REPORTS:
        for rr, e in zip(raw[n], aligned[n]):
            key_text = raw_canonical(rr)
            if not has_cyrillic(key_text):
                continue
            key = normalize_description(key_text)
            en = normalize_description(eng_romanized(e))
            if not en or en == key:
                continue
            entry = dict_map.setdefault(key, {"en": Counter(), "occ": 0, "bg_sample": key_text})
            entry["en"][en] += 1
            entry["occ"] += 1

    translation_rows = [{
        "normalized_key": k,
        "bg_sample": v["bg_sample"],
        "en_translation": v["en"].most_common(1)[0][0],
        "occurrences": v["occ"],
        "source": "mined",
    } for k, v in dict_map.items()]
    translation_rows.sort(key=lambda r: (-r["occurrences"], r["normalized_key"]))

    # ---- (b) Vendor rules ----
    orderer_names = set()
    for n in (1, 3):
        for rr in raw[n]:
            if rr["orderer"]:
                orderer_names.add(rr["orderer"].lower())

    # (b1) PRIMARY: RAW Cyrillic beneficiary -> category
    raw_vendor_cat = defaultdict(Counter)
    raw_vendor_pref = defaultdict(Counter)
    for n in (1, 3):   # report(2) has no counterparty column
        for rr, e in zip(raw[n], aligned[n]):
            ben = rr["beneficiary"]
            if not ben or ben.lower() in orderer_names:
                continue
            cat = eng_category(e)
            if cat not in CATEGORY_IDS:
                if cat:
                    unknown_categories[cat] += 1
                continue
            raw_vendor_cat[ben][cat] += 1
            if n == PREFERRED_REPORT:
                raw_vendor_pref[ben][cat] += 1

    # (b2) SECONDARY: romanized english Vendor -> category (future xlsx uploads)
    rom_vendor_cat = defaultdict(Counter)
    rom_vendor_pref = defaultdict(Counter)
    for n in REPORTS:
        for e in english[n]:
            vend = clean(e.get("Vendor"))
            if not vend or vend.lower() == "unknown vendor":
                continue
            cat = eng_category(e)
            if cat not in CATEGORY_IDS:
                if cat:
                    unknown_categories[cat] += 1
                continue
            rom_vendor_cat[vend][cat] += 1
            if n == PREFERRED_REPORT:
                rom_vendor_pref[vend][cat] += 1

    def majority(counter, pref):
        top = counter.most_common()
        best = top[0][1]
        tied = [c for c, n in top if n == best]
        if len(tied) == 1:
            return tied[0]
        cand = [c for c in tied if pref.get(c)]
        if cand:
            return max(cand, key=lambda c: pref[c])
        return sorted(tied)[0]

    def build_vendor_rows(cat_map, pref_map, priority, provenance):
        rows = []
        dropped_fallback = 0
        for pattern, counter in cat_map.items():
            chosen = majority(counter, pref_map.get(pattern, Counter()))
            cid = CATEGORY_IDS[chosen]
            if cid == FALLBACK_CATEGORY_ID:      # F5: never seed fallback rules
                dropped_fallback += 1
                continue
            rows.append({
                "match_type": "exact", "pattern": pattern, "category_id": cid,
                "category_name": chosen, "priority": priority,
                "hits_in_reference": sum(counter.values()), "provenance": provenance,
            })
        rows.sort(key=lambda r: (-r["hits_in_reference"], r["pattern"].lower()))
        return rows, dropped_fallback

    raw_rows, raw_dropped = build_vendor_rows(raw_vendor_cat, raw_vendor_pref,
                                              VENDOR_PRIORITY_RAW, "raw_cyrillic")
    rom_rows, rom_dropped = build_vendor_rows(rom_vendor_cat, rom_vendor_pref,
                                              VENDOR_PRIORITY_ROMANIZED, "romanized_en")
    raw_patterns = {r["pattern"] for r in raw_rows}
    rom_rows = [r for r in rom_rows if r["pattern"] not in raw_patterns]
    vendor_rows = raw_rows + rom_rows

    # ---- (c) Keyword rules (validate against RAW canonical text) ----
    # Auto-derive merchant keyword candidates from the romanized Vendor names:
    # card/POS transactions have no counterparty field, so their merchant appears
    # only inside the raw description text (e.g. "...EMAILOCTOPUS/LONDON..."). A
    # description-substring keyword is the only thing that can categorize them at
    # ingest. Names that are transliterated Cyrillic people/companies (which live
    # in the beneficiary field, not the description) simply score 0 hits below and
    # are dropped — so this naturally keeps only Latin card-merchant tokens.
    auto_candidates = []
    seen_kw = {kw.upper() for kw, _c, _p in KEYWORD_CANDIDATES}
    for vend, counter in rom_vendor_cat.items():
        chosen = majority(counter, rom_vendor_pref.get(vend, Counter()))
        cid = CATEGORY_IDS[chosen]
        vu = vend.upper()
        if cid == FALLBACK_CATEGORY_ID or len(vend.strip()) < 4 or vu in seen_kw:
            continue
        seen_kw.add(vu)
        auto_candidates.append((vend, cid, 15))   # merchant priority between 10 and 20

    keyword_rows = []
    keyword_report = {}
    for kw, cat_id, priority in KEYWORD_CANDIDATES + auto_candidates:
        kw_l = kw.lower()
        cc = Counter()
        for n in REPORTS:
            for rr, e in zip(raw[n], aligned[n]):
                if kw_l in raw_canonical(rr).lower():
                    cc[eng_category(e)] += 1
        total = sum(cc.values())
        if total == 0:
            keyword_report[kw] = "SKIP no-hits"
            continue
        dom_cat, dom_n = cc.most_common(1)[0]
        dominance = dom_n / total
        assigned = ID_TO_NAME[cat_id]
        if total >= MIN_KEYWORD_HITS and dominance >= MIN_KEYWORD_DOMINANCE and dom_cat == assigned \
           and cat_id != FALLBACK_CATEGORY_ID:
            keyword_rows.append({"keyword": kw, "category_id": cat_id, "category_name": assigned,
                                 "priority": priority, "hits_in_reference": total,
                                 "force_review": kw in FORCE_REVIEW_KEYWORDS})
            keyword_report[kw] = f"OK dom={dominance:.0%} n={total}"
        else:
            keyword_report[kw] = f"SKIP dom_cat={dom_cat!r} dom={dominance:.0%} n={total}"
    keyword_rows.sort(key=lambda r: (r["priority"], -r["hits_in_reference"], r["keyword"]))

    # ---- (d) Overhead type per category (report(3) dominant; fallback -> NULL) ----
    overhead_counter = defaultdict(Counter)
    for e in english[PREFERRED_REPORT]:
        cat = eng_category(e)
        if cat in CATEGORY_IDS:
            overhead_counter[cat][clean(e.get("Overhead Type")) or None] += 1
    for cat in CATEGORY_IDS:
        if cat not in overhead_counter:
            for n in REPORTS:
                for e in english[n]:
                    if eng_category(e) == cat:
                        overhead_counter[cat][clean(e.get("Overhead Type")) or None] += 1
    overhead_by_category, overhead_detail = {}, {}
    for cat, cid in CATEGORY_IDS.items():
        counter = overhead_counter.get(cat, Counter())
        total = sum(counter.values())
        chosen = None
        if total:
            dv, dn = counter.most_common(1)[0]
            if dv in ("Fixed", "Variable") and dn / total >= OVERHEAD_DOMINANCE:
                chosen = dv
        if cid == FALLBACK_CATEGORY_ID:
            chosen = None
        overhead_by_category[cat] = chosen
        overhead_detail[cat] = {"chosen": chosen, "counts": dict(counter), "total": total}

    # ---- (e) SELF-VALIDATION: replicate ingest matching over ALL raw rows ----
    vendor_lookup = {}
    for r in sorted(vendor_rows, key=lambda x: x["priority"]):
        vendor_lookup.setdefault(r["pattern"].lower(), r["category_id"])
    kw_sorted = sorted(keyword_rows, key=lambda x: x["priority"])
    dict_keys = {r["normalized_key"] for r in translation_rows}

    def categorize(rr):
        ben = rr["beneficiary"].lower()
        if ben and ben in vendor_lookup:
            return vendor_lookup[ben], "vendor_rule"
        canon = raw_canonical(rr).lower()
        for r in kw_sorted:
            if r["keyword"].lower() in canon:
                return r["category_id"], "keyword_rule"
        return FALLBACK_CATEGORY_ID, "fallback"

    dict_hits = dict_total = 0
    agree = compared = predicted_misc = total_rows = 0
    per_report_val = {}
    for n in REPORTS:
        rh = rt = ag = cmp = pm = rows_n = 0
        for rr, e in zip(raw[n], aligned[n]):
            rows_n += 1
            total_rows += 1
            canon = raw_canonical(rr)
            if has_cyrillic(canon):
                rt += 1
                if normalize_description(canon) in dict_keys:
                    rh += 1
            cid, _src = categorize(rr)
            if cid == FALLBACK_CATEGORY_ID:
                pm += 1
            truth = eng_category(e)
            if truth in CATEGORY_IDS:
                cmp += 1
                if CATEGORY_IDS[truth] == cid:
                    ag += 1
        per_report_val[n] = {
            "rows": rows_n,
            "dict_hit_rate": round(rh / rt, 4) if rt else None, "cyrillic_rows": rt,
            "category_agreement": round(ag / cmp, 4) if cmp else None, "compared": cmp,
            "predicted_misc": pm, "predicted_misc_share": round(pm / rows_n, 4) if rows_n else None,
        }
        dict_hits += rh; dict_total += rt; agree += ag; compared += cmp; predicted_misc += pm

    ref_misc = {n: sum(1 for e in english[n] if eng_category(e) == "Miscellaneous") for n in REPORTS}

    validation = {
        "overall_dict_hit_rate": round(dict_hits / dict_total, 4) if dict_total else None,
        "overall_dict_cyrillic_rows": dict_total,
        "overall_category_agreement": round(agree / compared, 4) if compared else None,
        "overall_compared": compared,
        "overall_predicted_misc": predicted_misc,
        "overall_predicted_misc_share": round(predicted_misc / total_rows, 4) if total_rows else None,
        "reference_misc_counts": ref_misc,
        "reference_misc_share_report3": round(ref_misc[3] / len(english[3]), 4),
        "per_report": per_report_val,
    }

    stats = {
        "align_stats": align_stats,
        "orderer_names_excluded": sorted(orderer_names),
        "vendor_rules_raw": len(raw_rows), "vendor_rules_romanized": len(rom_rows),
        "vendor_dropped_fallback_raw": raw_dropped, "vendor_dropped_fallback_romanized": rom_dropped,
        "keyword_validation": keyword_report,
        "unknown_categories": dict(unknown_categories),
        "overhead_detail": overhead_detail,
        "validation": validation,
    }
    return {
        "translation_rows": translation_rows,
        "vendor_rows": vendor_rows,
        "keyword_rows": keyword_rows,
        "overhead_by_category": overhead_by_category,
        "stats": stats,
    }


# ---------------------------------------------------------------------------
# Emitters
# ---------------------------------------------------------------------------

def q(value):
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def write_csv(path, rows, fields):
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_json(path, obj):
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(obj, fh, ensure_ascii=False, indent=2)


def emit_seed_sql(result):
    t, v, k = result["translation_rows"], result["vendor_rows"], result["keyword_rows"]
    st = result["stats"]
    L = []
    a = L.append
    a("-- ============================================================================")
    a("-- Migration 127: Expense Mapping Seed (GENERATED — do not hand-edit)")
    a("-- ============================================================================")
    a("-- Generated by scripts/expenses/mine_expenses.py from the B Team 2025-AP RAW")
    a("-- bank exports (read-only), per CANONICAL CONTRACT v2. Ingest matches raw")
    a("-- Cyrillic rows, so all keys/patterns are the RAW Cyrillic fields.")
    a("--   - expense_translation_dict  key = normalize(Основание за плащане + ' ' +")
    a("--                               Описание на операцията); en = romanized form")
    a("--   - expense_vendor_rules       pattern = RAW Cyrillic Бенефициент (priority")
    a(f"--                               {VENDOR_PRIORITY_RAW}); romanized english Vendor kept as")
    a(f"--                               secondary (priority {VENDOR_PRIORITY_ROMANIZED}). No category-15 rules.")
    a("--   - expense_keyword_rules      re-validated on RAW description text")
    a("--")
    a("-- Self-validation on the reference corpus:")
    a(f"--   dictionary hit rate (Cyrillic rows) = {st['validation']['overall_dict_hit_rate']}")
    a(f"--   vendor+keyword category agreement    = {st['validation']['overall_category_agreement']}")
    a(f"--   predicted Miscellaneous share        = {st['validation']['overall_predicted_misc_share']}")
    a(f"-- Counts: {len(t)} translation entries, {len(v)} vendor rules "
      f"({st['vendor_rules_raw']} raw-cyr + {st['vendor_rules_romanized']} romanized), {len(k)} keyword rules.")
    a("-- Idempotent (ON CONFLICT DO NOTHING); Cyrillic/apostrophes escaped via doubled quotes.")
    a("-- ============================================================================")
    a("")
    a("BEGIN;")
    a("")
    a("-- ----------------------------------------------------------------------------")
    a("-- expense_translation_dict")
    a("-- ----------------------------------------------------------------------------")
    if t:
        a("INSERT INTO public.expense_translation_dict")
        a("    (normalized_key, bg_sample, en_translation, occurrences, source)")
        a("VALUES")
        last = len(t) - 1
        for i, r in enumerate(t):
            sep = "" if i == last else ","
            a(f"    ({q(r['normalized_key'])}, {q(r['bg_sample'])}, {q(r['en_translation'])}, "
              f"{q(r['occurrences'])}, {q(r['source'])}){sep}")
        a("ON CONFLICT (normalized_key) DO NOTHING;")
    else:
        a("-- (no translation entries mined)")
    a("")
    a("-- ----------------------------------------------------------------------------")
    a("-- expense_vendor_rules  (raw_cyrillic = primary; romanized_en = secondary)")
    a("-- ----------------------------------------------------------------------------")
    if v:
        a("INSERT INTO public.expense_vendor_rules")
        a("    (match_type, pattern, category_id, priority, hits_in_reference, source)")
        a("VALUES")
        last = len(v) - 1
        for i, r in enumerate(v):
            sep = "" if i == last else ","
            a(f"    ({q(r['match_type'])}, {q(r['pattern'])}, {q(r['category_id'])}, "
              f"{q(r['priority'])}, {q(r['hits_in_reference'])}, {q(r['provenance'])}){sep}  -- {r['category_name']}")
        a("ON CONFLICT (match_type, pattern) DO NOTHING;")
    else:
        a("-- (no vendor rules mined)")
    a("")
    a("-- ----------------------------------------------------------------------------")
    a("-- expense_keyword_rules")
    a("-- ----------------------------------------------------------------------------")
    if k:
        a("INSERT INTO public.expense_keyword_rules")
        a("    (keyword, category_id, priority, hits_in_reference, force_review)")
        a("VALUES")
        last = len(k) - 1
        for i, r in enumerate(k):
            sep = "" if i == last else ","
            flag = "  -- FORCE REVIEW: broad bank-name substring" if r["force_review"] else ""
            a(f"    ({q(r['keyword'])}, {q(r['category_id'])}, {q(r['priority'])}, "
              f"{q(r['hits_in_reference'])}, {q(r['force_review'])}){sep}  -- {r['category_name']}{flag}")
        a("ON CONFLICT (keyword) DO NOTHING;")
    else:
        a("-- (no keyword rules mined)")
    a("")
    a("COMMIT;")
    a("")
    with open(SEED_SQL_PATH, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L))


# ---------------------------------------------------------------------------
# DELTA mining (migration 128) — validate additive keyword rules only.
# ---------------------------------------------------------------------------

def mine_delta(reference_dir):
    """Validate DELTA_KEYWORD_CANDIDATES against the reference corpus using the
    exact same bar as the seed, and attach the documented exceptions. Returns
    (rows, report). Does NOT regenerate any seed artifact."""
    raw = {
        1: parse_report1(os.path.join(reference_dir, "report(1).xls.xlsx")),
        2: parse_report2(os.path.join(reference_dir, "report(2).xls.xlsx")),
        3: parse_report3_html(os.path.join(reference_dir, "report(3).xls")),
    }
    english = {n: load_english(os.path.join(reference_dir, f"report({n})_english.xlsx")) for n in REPORTS}
    aligned = {n: align(raw[n], english[n])[0] for n in REPORTS}

    def raw_canonical(rr):
        return (rr["osnovanie"] + " " + rr["opisanie"]).strip()

    rows, report = [], {}
    for kw, cat_id, priority in DELTA_KEYWORD_CANDIDATES:
        kw_l = kw.lower()
        cc = Counter()
        for n in REPORTS:
            for rr, e in zip(raw[n], aligned[n]):
                if kw_l in raw_canonical(rr).lower():
                    cc[eng_category(e)] += 1
        total = sum(cc.values())
        assigned = ID_TO_NAME[cat_id]
        if total == 0:
            report[kw] = "SKIP no-hits"
            continue
        dom_cat, dom_n = cc.most_common(1)[0]
        dominance = dom_n / total
        if total >= MIN_KEYWORD_HITS and dominance >= MIN_KEYWORD_DOMINANCE \
           and dom_cat == assigned and cat_id != FALLBACK_CATEGORY_ID:
            rows.append({"keyword": kw, "category_id": cat_id, "category_name": assigned,
                         "priority": priority, "hits_in_reference": total,
                         "note": "validated (>=3 hits, >=90% dominance)"})
            report[kw] = f"OK dom={dominance:.0%} n={total}"
        else:
            report[kw] = f"SKIP dom_cat={dom_cat!r} dom={dominance:.0%} n={total}"

    for kw, cat_id, priority, justification in DELTA_KEYWORD_EXCEPTIONS:
        rows.append({"keyword": kw, "category_id": cat_id, "category_name": ID_TO_NAME[cat_id],
                     "priority": priority, "hits_in_reference": 0,
                     "note": f"EXCEPTION: {justification}"})
        report[kw] = "EXCEPTION (documented, bypasses reference bar)"

    rows.sort(key=lambda r: (r["priority"], r["keyword"]))
    return rows, report


def emit_delta_sql(rows):
    L = []
    a = L.append
    a("-- ============================================================================")
    a("-- Migration 128: Expense Mapping DELTA (GENERATED — do not hand-edit)")
    a("-- ============================================================================")
    a("-- Generated by `scripts/expenses/mine_expenses.py --delta`. Additive-only:")
    a("-- migration 127 is already applied to prod and is NOT regenerated; this file")
    a("-- is a strict idempotent SUPERSET (any already-applied row is a no-op via ON")
    a("-- CONFLICT). Two root causes are addressed, plus business-owner directives:")
    a("--   (1) merchant-descriptor variants: the seed's clean-vendor-name keywords")
    a("--       ('HIGHLEVEL INC.', 'KAFFEKAPSLEN', 'EPAY OFFICE1.BG') are not")
    a("--       substrings of the shorter POS descriptors that appear in real card")
    a("--       lines, so those rows fell through to Miscellaneous.")
    a("--   (2) English keyword candidates ('RECURRING FEE') were mined against")
    a("--       CYRILLIC raw text -> 0 hits -> dropped; and categorize() matches")
    a("--       description_original (Cyrillic), never the English translation. The")
    a("--       reference-backed CYRILLIC anchors below are the effective fee rules")
    a("--       ('ТАКСА ЗА ИЗХ.ПРЕВОД SEPA' 23/23, 'ПЕРИОДИЧНА ТАКСА' 25/25 Bank Fees).")
    a("--       Anchored on the 'ТАКСА ЗА' (fee-for) prefix so they do NOT touch the")
    a("--       145 salary/tax/insurance rows that share the bare 'ИЗХ.ПРЕВОД SEPA'.")
    a("-- Validated rows meet the seed bar (>=3 reference hits, >=90% dominance, a")
    a("-- non-15 dominant category); exception rows carry an inline justification")
    a("-- (owner-named English literals are inert on Cyrillic prod, future-proofing).")
    a("-- Idempotent: ON CONFLICT (keyword) DO NOTHING. No vendor->category-15 rules.")
    a("-- ============================================================================")
    a("")
    a("BEGIN;")
    a("")
    a("INSERT INTO public.expense_keyword_rules")
    a("    (keyword, category_id, priority, hits_in_reference, force_review)")
    a("VALUES")
    last = len(rows) - 1
    for i, r in enumerate(rows):
        sep = "" if i == last else ","
        a(f"    ({q(r['keyword'])}, {q(r['category_id'])}, {q(r['priority'])}, "
          f"{q(r['hits_in_reference'])}, FALSE){sep}  -- {r['category_name']} :: {r['note']}")
    a("ON CONFLICT (keyword) DO NOTHING;")
    a("")
    a("COMMIT;")
    a("")
    with open(DELTA_SQL_PATH, "w", encoding="utf-8") as fh:
        fh.write("\n".join(L))


def run_delta():
    reference_dir = _resolve_reference_dir()
    print(f"Reference dir: {reference_dir}")
    rows, report = mine_delta(reference_dir)
    emit_delta_sql(rows)
    print("\n=== DELTA KEYWORD VALIDATION ===")
    for kw, status in report.items():
        print(f"  {kw:<14} {status}")
    print(f"\nEmitted {len(rows)} delta keyword rule(s) -> {DELTA_SQL_PATH}")


def _resolve_reference_dir():
    for d in CANDIDATE_REFERENCE_DIRS:
        if os.path.isdir(d):
            return d
    sys.exit("ERROR: reference folder not found in any known Dropbox mount.")


def main():
    if "--delta" in sys.argv:
        run_delta()
        return
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    reference_dir = _resolve_reference_dir()
    print(f"Reference dir: {reference_dir}")

    result = mine(reference_dir)
    st = result["stats"]

    write_csv(os.path.join(OUTPUT_DIR, "translation_dict.csv"), result["translation_rows"],
              ["normalized_key", "bg_sample", "en_translation", "occurrences", "source"])
    write_json(os.path.join(OUTPUT_DIR, "translation_dict.json"), result["translation_rows"])
    write_csv(os.path.join(OUTPUT_DIR, "vendor_rules.csv"), result["vendor_rows"],
              ["match_type", "pattern", "category_id", "category_name", "priority", "hits_in_reference", "provenance"])
    write_json(os.path.join(OUTPUT_DIR, "vendor_rules.json"), result["vendor_rows"])
    write_csv(os.path.join(OUTPUT_DIR, "keyword_rules.csv"), result["keyword_rows"],
              ["keyword", "category_id", "category_name", "priority", "hits_in_reference", "force_review"])
    write_json(os.path.join(OUTPUT_DIR, "keyword_rules.json"), result["keyword_rows"])
    write_json(os.path.join(OUTPUT_DIR, "overhead_by_category.json"), result["overhead_by_category"])
    write_json(os.path.join(OUTPUT_DIR, "mining_stats.json"), st)
    emit_seed_sql(result)

    val = st["validation"]
    print("\n=== ALIGNMENT ===")
    for n in REPORTS:
        s = st["align_stats"][n]
        print(f"  report{n}: raw={s['raw']} eng={s['english']} by_key={s['by_key']} "
              f"by_pos={s['by_pos']} unaligned={s['unaligned']}")
    print("\n=== SEED STATS ===")
    print(f"  translation entries : {len(result['translation_rows'])}")
    print(f"  vendor rules        : {len(result['vendor_rows'])} "
          f"(raw-cyr={st['vendor_rules_raw']}, romanized={st['vendor_rules_romanized']}; "
          f"dropped cat15 raw={st['vendor_dropped_fallback_raw']}, rom={st['vendor_dropped_fallback_romanized']})")
    print(f"  keyword rules       : {len(result['keyword_rows'])}")
    print(f"  unknown categories  : {st['unknown_categories'] or 'none'}")
    print("\n=== SELF-VALIDATION ===")
    print(f"  dict hit rate (Cyrillic rows) : {val['overall_dict_hit_rate']} "
          f"over {val['overall_dict_cyrillic_rows']} rows  (target >0.90)")
    print(f"  category agreement            : {val['overall_category_agreement']} over {val['overall_compared']} rows")
    print(f"  predicted Misc share          : {val['overall_predicted_misc_share']} "
          f"(reference report3 = {val['reference_misc_share_report3']})")
    for n in REPORTS:
        pr = val["per_report"][n]
        print(f"    report{n}: dict={pr['dict_hit_rate']} agree={pr['category_agreement']} "
              f"misc_share={pr['predicted_misc_share']}")
    print("\n=== KEYWORD VALIDATION ===")
    for kw, status in st["keyword_validation"].items():
        print(f"  {kw:<28} {status}")
    print(f"\nSeed SQL: {SEED_SQL_PATH}")


if __name__ == "__main__":
    main()
